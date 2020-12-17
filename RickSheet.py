import openpyxl
# Assume Rick won't input a zero anywhere

def parts(requests, LC):
    loc_name = input("Enter the location name: ")
    requests[loc_name] = []

    parts = input("Enter in all the parts you want sent to this location, separate all part numbers with a space\n(Only press the enter key when your list is complete)\n")
    parts = parts.split()

    for index, part in enumerate(parts):
        requests[loc_name].append([part])
        while(not LC):
            qty = input(f"How much for {part}: ")
            if not qty.isdigit():
                print("Error: Your input wasn't a digit or was negative\nTry again")
            elif int(qty) is 0:
                print(f"Error: You cannot request zero {part}\nTry again")
            else:
                requests[loc_name][index].append(qty)
                LC = True
        LC = False

def excel(requests):
    wb_obj = openpyxl.Workbook()  # Create a new excel workbook
    sheet_obj = wb_obj.active  # Get the current (only) sheet and store it into an object

    headers = ["Part No", "Qty", "Location", "Status", "Request From"]
    for index in range(2, 7):
        sheet_obj.cell(row=1, column=index).value = headers[index-2]

    r_ow = 2
    for key in requests:
        for index in range(len(requests[key])):
            sheet_obj.cell(row=r_ow, column=1).value = "Rick Mally Sr."
            sheet_obj.cell(row=r_ow, column=2).value = requests[key][index][0]
            sheet_obj.cell(row=r_ow, column=3).value = requests[key][index][1]
            sheet_obj.cell(row=r_ow, column=4).value = key
            sheet_obj.cell(row=r_ow, column=5).value = "REQUESTED"
            r_ow += 1
            # Column 6 data left blank so the default label will set in after pulse transfer

    wb_obj.save("C:\\Users\\andre.le\\Downloads\\Sample Request.xlsx")

def main():
    print("Welcome Rick Sr.")

    request_dict = {}  # Key is Location, Value is list pair of Part No and Qty
    loop_cond = False
    while(not loop_cond):
        locs = input("Enter in the amount of locations you want parts sent to: ")
        if not locs.isdigit():
            print("Error: Your input wasn't a digit or was negative\nTry again")
        elif int(locs) is 0:
            print("You have indicated that you wish to send parts to zero locations\nEnding program")
            break
        else:
            locs = int(locs)
            loop_cond = True

    if loop_cond is False:
        pass
    else:
        loop_cond = False
        for index in range(locs):
            print("For location: % 2d" %(index+1))
            parts(request_dict, loop_cond)
        excel(request_dict)

main()
